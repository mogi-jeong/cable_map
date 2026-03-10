"""
make_application.py
이음 v2 추출 xlsx + OH_FLD_INVS xlsx → 공가 신청서 xlsx (3시트)

시트 구성:
  시설계획서_신규  : 장표 없는 전주 → 접수2행
  시설계획서_정비  : [신설] 접수2행 전체 묶음 → [해제] 접수3행 전체 묶음
  시설계획서_해지  : 자가주 해지 (추후)

사용법:
  python make_application.py \
    --invs   OH_FLD_INVS_4335202603100007.xlsx \
    --poles  전주_숙암지1_신월간-297.xlsx \
    --out    공가신청서_신월간297.xlsx
"""

import argparse
import re
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_ROW1 = [
    '순번','접수구분',
    '현전주','현전주','현전주','현전주',
    '1차전주','1차전주','1차전주','1차전주',
    '2차전주','2차전주','2차전주','2차전주',
    '통신케이블','통신케이블','통신케이블','통신케이블','통신케이블',
    '통신케이블','통신케이블','통신케이블','통신케이블','통신케이블',
    '통신기기1','통신기기1','통신기기1',
    '통신기기2','통신기기2','통신기기2',
    '통신기기3','통신기기3','통신기기3',
    '비고',
]
HEADER_ROW2 = [
    '순번','접수구분',
    '선로명','선로번호','관리구','번호',
    '선로명','선로번호','관리구','번호',
    '선로명','선로번호','관리구','번호',
    '설치단','사업자','설치일자','케이블번호','용도',
    '통선선종류','규격','승인코드','고객공급선종류','봉인번호',
    '기기코드','사업자','관리번호',
    '기기코드','사업자','관리번호',
    '기기코드','사업자','관리번호',
    '비고',
]
NCOLS = len(HEADER_ROW2)

C = {
    '순번':0,'접수구분':1,
    '현선로명':2,'현선로번호':3,'현관리구':4,'현번호':5,
    '1차선로명':6,'1차선로번호':7,'1차관리구':8,'1차번호':9,
    '2차선로명':10,'2차선로번호':11,'2차관리구':12,'2차번호':13,
    '설치단':14,'사업자':15,'설치일자':16,'케이블번호':17,
    '용도':18,'통선선종류':19,'규격':20,'승인코드':21,
    '고객공급선종류':22,'봉인번호':23,
    '비고':33,
}

JUMP_THRESHOLD = 2

COLOR = {
    'header':  'D9E1F2',
    'sep_신설': 'BDD7EE',
    'sep_해제': 'FCE4D6',
    'row2':    'E2EFDA',
    'row3':    'FFF2CC',
}

COL_WIDTHS = {
    1:6, 2:7,
    3:9, 4:8, 5:8, 6:5,
    7:9, 8:8, 9:8, 10:5,
    11:9, 12:8, 13:8, 14:5,
    15:5, 16:13, 17:11, 18:18,
    19:5, 20:5, 21:5, 22:7, 23:7, 24:8,
    34:20,
}


def v(x):
    if x is None: return ''
    try:
        import math
        if math.isnan(float(x)): return ''
    except (TypeError, ValueError): pass
    s = str(x).strip()
    return '' if s.lower() in ('nan','none') else s


def make_id(관리구, 번호):
    try:
        n = str(int(str(번호).strip())).zfill(3)
    except (ValueError, TypeError):
        n = str(번호).strip()
    return (str(관리구).strip().upper() + n).upper()


def sort_key(pole):
    s = str(pole['선로번호'])
    has_branch = bool(re.search(r'[A-Za-z]', s))
    base_nums  = [int(p) for p in re.findall(r'\d+', s)]
    return (base_nums[0] if base_nums else 0, 0 if has_branch else 1, s)


def num_str(번호):
    try: return str(int(번호))
    except (ValueError, TypeError): return str(번호)


def load_poles(path):
    df = pd.read_excel(path, header=0, dtype=str)
    df.columns = ['관리구','번호','선로명','선로번호','경간']
    poles = []
    for _, row in df.iterrows():
        관리구 = v(row['관리구'])
        번호   = v(row['번호'])
        if not 관리구: continue
        poles.append({
            '관리구': 관리구, '번호': 번호,
            '전산화번호': make_id(관리구, 번호),
            '선로명': v(row['선로명']),
            '선로번호': v(row['선로번호']),
            '경간': v(row['경간']),
        })
    return poles


def load_invs(path):
    df = pd.read_excel(path, dtype=str)
    df.columns = df.columns.str.strip()
    df = df.dropna(subset=['시작전산화번호'])
    df['시작전산화번호'] = df['시작전산화번호'].str.strip().str.upper()
    by_id, by_name = {}, {}
    for _, row in df.iterrows():
        r = row.to_dict()
        k = r['시작전산화번호']
        by_id.setdefault(k, []).append(r)
        nm = v(r.get('현전주 선로명',''))
        nb = v(r.get('현전주 선로번호',''))
        if nm:
            by_name.setdefault((nm, nb), []).append(r)
    return by_id, by_name


def invs_to_row3(r, seq, note=''):
    row = ['']*NCOLS
    row[C['순번']]        = seq
    row[C['접수구분']]    = '3'
    row[C['현선로명']]    = v(r.get('현전주 선로명'))
    row[C['현선로번호']]  = v(r.get('현전주 선로번호'))
    row[C['현관리구']]    = v(r.get('현전주 관리구'))
    row[C['현번호']]      = v(r.get('현전주 번호'))
    row[C['1차선로명']]   = v(r.get('1차전주 선로명'))
    row[C['1차선로번호']] = v(r.get('1차전주 선로번호'))
    row[C['1차관리구']]   = v(r.get('1차전주 관리구'))
    row[C['1차번호']]     = v(r.get('1차전주 번호'))
    row[C['2차선로명']]   = v(r.get('2차전주 선로명'))
    row[C['2차선로번호']] = v(r.get('2차전주 선로번호'))
    row[C['2차관리구']]   = v(r.get('2차전주 관리구'))
    row[C['2차번호']]     = v(r.get('2차전주 번호'))
    row[C['설치단']]      = v(r.get('설치단'))
    row[C['사업자']]      = v(r.get('사업자'))
    row[C['설치일자']]    = v(r.get('설치일자'))
    row[C['케이블번호']]  = v(r.get('케이블번호'))
    row[C['용도']]        = v(r.get('용도'))
    row[C['통선선종류']]  = v(r.get('통신선종류'))
    row[C['규격']]        = v(r.get('규격'))
    row[C['승인코드']]    = v(r.get('승인코드'))
    row[C['고객공급선종류']] = v(r.get('고객공급선종류'))
    row[C['봉인번호']]    = v(r.get('봉인번호'))
    row[C['비고']]        = note
    return row


def pole_to_row2(pole, prev, nxt, seq, defaults, invs_row, note=''):
    row = ['']*NCOLS
    row[C['순번']]        = seq
    row[C['접수구분']]    = '2'
    row[C['현선로명']]    = pole['선로명']
    row[C['현선로번호']]  = pole['선로번호']
    row[C['현관리구']]    = pole['관리구']
    row[C['현번호']]      = num_str(pole['번호'])
    if prev:
        row[C['1차선로명']]   = prev['선로명']
        row[C['1차선로번호']] = prev['선로번호']
        row[C['1차관리구']]   = prev['관리구']
        row[C['1차번호']]     = num_str(prev['번호'])
    else:
        row[C['1차관리구']] = '99999'; row[C['1차번호']] = '999'
    if nxt:
        row[C['2차선로명']]   = nxt['선로명']
        row[C['2차선로번호']] = nxt['선로번호']
        row[C['2차관리구']]   = nxt['관리구']
        row[C['2차번호']]     = num_str(nxt['번호'])
    else:
        row[C['2차관리구']] = '99999'; row[C['2차번호']] = '999'
    row[C['설치단']]     = defaults['설치단']
    row[C['사업자']]     = defaults['사업자']
    row[C['용도']]       = defaults['용도']
    row[C['통선선종류']] = v(invs_row.get('통신선종류')) if invs_row else defaults['통신선']
    row[C['규격']]       = v(invs_row.get('규격'))       if invs_row else defaults['규격']
    row[C['비고']]       = note
    return row


def check_jaga(pole, prev, nxt):
    notes = []
    try:
        cur = int(pole['선로번호'])
        if prev and abs(cur - int(prev['선로번호'])) > JUMP_THRESHOLD:
            notes.append('1차 전주 자가주')
        if nxt and abs(int(nxt['선로번호']) - cur) > JUMP_THRESHOLD:
            notes.append('2차 전주 자가주')
    except (ValueError, TypeError): pass
    return ' / '.join(notes)


def classify(poles, by_id, by_name, defaults):
    신규, 정비_신설, 정비_해제 = [], [], []
    seq = 1
    for i, pole in enumerate(poles):
        prev = poles[i-1] if i > 0 else None
        nxt  = poles[i+1] if i < len(poles)-1 else None
        invs_rows  = by_id.get(pole['전산화번호'])
        match_type = 'id'
        if not invs_rows:
            invs_rows  = by_name.get((pole['선로명'], pole['선로번호']))
            match_type = 'name' if invs_rows else None
        jaga = check_jaga(pole, prev, nxt)

        if invs_rows:
            print(f"  [정비] {pole['전산화번호']:10s} {pole['선로명']}-{pole['선로번호']:6s} "
                  f"match={match_type} 장표{len(invs_rows)}행" + (f" [{jaga}]" if jaga else ''))
            r2 = pole_to_row2(pole, prev, nxt, seq, defaults, invs_rows[0], jaga)
            정비_신설.append({'type':'row2','data':r2})
            for ir in invs_rows:
                r3 = invs_to_row3(ir, seq, jaga)
                정비_해제.append({'type':'row3','data':r3})
        else:
            print(f"  [신규] {pole['전산화번호']:10s} {pole['선로명']}-{pole['선로번호']:6s}"
                  + (f" [{jaga}]" if jaga else ''))
            r2 = pole_to_row2(pole, prev, nxt, seq, defaults, None, jaga)
            신규.append({'type':'row2','data':r2})
        seq += 1
    return 신규, 정비_신설, 정비_해제


def write_sheet(ws, data_rows, sheet_title):
    t = Side(style='thin')
    border = Border(left=t, right=t, top=t, bottom=t)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def sc(row, col, value='', fill=None, bold=False, size=9):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font      = Font(name='맑은 고딕', bold=bold, size=size)
        cell.fill      = PatternFill('solid', start_color=fill) if fill else PatternFill()
        cell.alignment = center
        cell.border    = border
        return cell

    # 제목
    ws.merge_cells(f'A1:{get_column_letter(NCOLS)}1')
    c = ws.cell(row=1, column=1, value=sheet_title)
    c.font = Font(name='맑은 고딕', bold=True, size=12)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    # 헤더
    for ci, val in enumerate(HEADER_ROW1, 1):
        sc(2, ci, val, fill=COLOR['header'], bold=True)
    for ci, val in enumerate(HEADER_ROW2, 1):
        sc(3, ci, val, fill=COLOR['header'], bold=True)
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18

    cur = 4
    for item in data_rows:
        if item['type'] == 'sep':
            ws.merge_cells(f'A{cur}:{get_column_letter(NCOLS)}{cur}')
            c = ws.cell(row=cur, column=1, value=f'▶  {item["label"]}')
            c.font      = Font(name='맑은 고딕', bold=True, size=10)
            c.fill      = PatternFill('solid', start_color=item['color'])
            c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            for ci in range(2, NCOLS+1):
                ws.cell(row=cur, column=ci).fill = PatternFill('solid', start_color=item['color'])
            ws.row_dimensions[cur].height = 16
        else:
            fill_c = COLOR['row2'] if item['type'] == 'row2' else COLOR['row3']
            for ci, val in enumerate(item['data'], 1):
                sc(cur, ci, val, fill=fill_c)
            ws.row_dimensions[cur].height = 15
        cur += 1

    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = 'A4'


def make_excel(신규, 정비_신설, 정비_해제, out_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet('시설계획서_신규')
    write_sheet(ws1, 신규, '공가 가공설비 시설계획서 (신규)')

    ws2 = wb.create_sheet('시설계획서_정비')
    정비 = (
        [{'type':'sep','label':'해제  (접수구분 3)','color':COLOR['sep_해제']}]
        + 정비_해제
        + [{'type':'sep','label':'신설  (접수구분 2)','color':COLOR['sep_신설']}]
        + 정비_신설
    )
    write_sheet(ws2, 정비, '공가 가공설비 시설계획서 (정비)')

    ws3 = wb.create_sheet('시설계획서_해지')
    write_sheet(ws3, [], '공가 가공설비 시설계획서 (해지)')

    wb.save(out_path)
    print(f"\n── 요약 {'─'*40}")
    print(f"  신규 시트: {len(신규)}행")
    print(f"  정비 시트: 신설 {len(정비_신설)}행 + 해제 {len(정비_해제)}행")
    print(f"  해지 시트: 0행 (자가주 추후)")
    print(f"✅ 저장 → {out_path}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--invs',   required=True)
    ap.add_argument('--poles',  required=True)
    ap.add_argument('--out',    default='공가신청서.xlsx')
    ap.add_argument('--설치단', default='2')
    ap.add_argument('--사업자', default='A000042286')
    ap.add_argument('--용도',   default='4')
    ap.add_argument('--통신선', default='O')
    ap.add_argument('--규격',   default='12')
    args = ap.parse_args()

    defaults = {
        '설치단': args.설치단, '사업자': args.사업자,
        '용도':   args.용도,   '통신선': args.통신선, '규격': args.규격,
    }

    print(f"\n▶ 이음 추출: {args.poles}")
    poles = load_poles(Path(args.poles))
    print(f"   전주: {len(poles)}개")

    print(f"▶ 장표: {args.invs}")
    by_id, by_name = load_invs(Path(args.invs))
    print(f"   키: {len(by_id):,}개")

    poles = sorted(poles, key=sort_key)

    print(f"\n── 매칭 {'─'*40}")
    신규, 정비_신설, 정비_해제 = classify(poles, by_id, by_name, defaults)
    make_excel(신규, 정비_신설, 정비_해제, Path(args.out))


if __name__ == '__main__':
    main()
